import openpyxl

#
# Finds a value in an excel sheet.
# <params> [path], [code to lookup]

def find_val(path, code):
    file = path
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(1):
        for cell in row:
            if cell.value == code:
                val_u = ws.cell(row=cell.row, column=2).value
                #print("Ran help func")  #change column number for any cell value you want
                return val_u